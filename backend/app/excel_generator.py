"""Geracao da planilha Excel formatada a partir de um orcamento aprovado/rascunho.

Layout replica visualmente os modelos de proposta da Rigel Engenharia:
pagina 1 com o cabecalho, dados do cliente e tabela de categorias/itens;
pagina 2 com preco global, condicoes, disposicoes gerais fixas e dados bancarios.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from . import models, schemas
from .numero_por_extenso import valor_por_extenso

EMPRESA_NOME = "RIGEL ENGENHARIA E IMPERMEABILIZAÇÕES"
EMPRESA_CNPJ = "39.617.612/0001-63"
EMPRESA_COR_BARRA = "1F4E79"

BANCO = "Itaú"
AGENCIA = "6541"
CONTA = "24672-1"
PIX = f"CNPJ : {EMPRESA_CNPJ}"

RESPONSAVEL_NOME = "Igor Pitanga"
RESPONSAVEL_CARGO = "Administração"
RESPONSAVEL_CONTATO = "cel.: (21) 97953-3142 / igor.pitanga@rigelengenharia.com / adm@rigelengenharia.com"

DISPOSICOES_GERAIS = [
    "Para elaboração da presente proposta, tomamos por base as informações fornecidas por V. Sas.",
    "Qualquer modificação, a partir do início de nossos trabalhos, que acarrete alteração no escopo "
    "apresentado, será revisado e cobrado correspondente ao serviço modificado.",
    "O prazo de validade desta proposta é de 30 dias.",
    "Todos os serviços alvo desta proposta serão realizados em dias úteis e em horário comercial.",
    "Os prazos indicados são estimativas e poderão ser revistos por situações extraordinárias, "
    "alterações solicitadas ou outros, sendo comunicado previamente.",
    "Todos os funcionários portarão os EPI's necessários nas frentes de serviços, e NR's necessárias.",
    "Deverá ser disponibilizado pelo contratante um ponto de energia próximo ao local de trabalho, "
    "assim como um ponto de água para execução de massa e outros serviços necessários.",
    "Deverá ser destinado pela contratante um vestiário para troca de roupa e um refeitório para "
    "almoço dos funcionários da contratada.",
]

MESES_PT = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]

STATUS_LABEL = {
    models.COBRANCA_POR_CONTA_CONTRATANTE: "por conta da contratante",
    models.COBRANCA_INCLUSO: "INCLUSO",
}

LARGURAS = {"A": 5, "B": 58, "C": 11, "D": 8, "E": 13, "F": 16}

_BORDA_FINA = Side(style="thin", color="B7B7B7")
_BORDA_GROSSA = Side(style="medium", color="1F4E79")


def _data_por_extenso(data) -> str:
    return f"{data.day:02d} de {MESES_PT[data.month - 1]} de {data.year}"


def _fmt_moeda(valor: float) -> str:
    texto = f"{valor:,.2f}"
    texto = texto.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {texto}"


def _cabecalho_empresa(ws: Worksheet, ultima_coluna: str) -> int:
    ws.merge_cells(f"A1:{ultima_coluna}1")
    celula = ws["A1"]
    celula.value = f"{EMPRESA_NOME} - CNPJ {EMPRESA_CNPJ}"
    celula.font = Font(bold=True, size=11, color=EMPRESA_COR_BARRA)
    celula.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{ultima_coluna}2")
    ws["A2"].fill = PatternFill("solid", fgColor=EMPRESA_COR_BARRA)
    ws.row_dimensions[2].height = 4
    return 3


def _aplicar_larguras(ws: Worksheet) -> None:
    for col, largura in LARGURAS.items():
        ws.column_dimensions[col].width = largura


def _linha_rotulo_valor(ws: Worksheet, linha: int, rotulo: str, valor: str, negrito_rotulo: bool = True) -> int:
    ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=negrito_rotulo, size=10)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
    celula = ws.cell(row=linha, column=2, value=valor)
    celula.font = Font(size=10)
    celula.alignment = Alignment(horizontal="left", wrap_text=True)
    return linha + 1


def _montar_pagina_itens(ws: Worksheet, orcamento: models.Orcamento, totais: schemas.OrcamentoTotais) -> None:
    _aplicar_larguras(ws)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    linha = _cabecalho_empresa(ws, "F")
    linha += 1

    numero = orcamento.numero_proposta or "—"
    ws.cell(row=linha, column=1, value=f"PROPOSTA Nº {numero}").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    data_cel = ws.cell(row=linha, column=4, value=f"Rio de Janeiro, {_data_por_extenso(orcamento.data_proposta)}")
    ws.merge_cells(start_row=linha, start_column=4, end_row=linha, end_column=6)
    data_cel.alignment = Alignment(horizontal="right")
    data_cel.font = Font(size=10)
    linha += 2

    ws.cell(row=linha, column=1, value="A").font = Font(bold=True, size=10)
    linha += 1
    ws.cell(row=linha, column=1, value=orcamento.cliente_nome or "").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    linha += 1
    linha = _linha_rotulo_valor(ws, linha, "Att:", orcamento.cliente_att or "")
    linha = _linha_rotulo_valor(ws, linha, "E-mail:", orcamento.cliente_email or "")
    linha += 1
    linha = _linha_rotulo_valor(ws, linha, "REF:", orcamento.referencia or "")
    linha = _linha_rotulo_valor(ws, linha, "End:", orcamento.endereco_obra or "")
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    intro = ws.cell(
        row=linha,
        column=1,
        value="Prezados, apresentamos nossa proposta tecnico-comercial para execução dos serviços abaixo descritos:",
    )
    intro.font = Font(bold=True, size=10)
    intro.alignment = Alignment(wrap_text=True)
    linha += 2

    cabecalho_tabela = ["Item", "Descrição", "Quantidade", "Unidade", "Preço Unit.", "Preço Total"]
    for coluna, titulo in enumerate(cabecalho_tabela, start=1):
        celula = ws.cell(row=linha, column=coluna, value=titulo)
        celula.font = Font(bold=True, size=9, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=EMPRESA_COR_BARRA)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = Border(top=_BORDA_GROSSA, bottom=_BORDA_GROSSA)
    linha += 1

    for indice_categoria, categoria in enumerate(orcamento.categorias, start=1):
        ws.cell(row=linha, column=1, value=indice_categoria).font = Font(bold=True, size=10)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
        titulo_cel = ws.cell(row=linha, column=2, value=(categoria.titulo or "").upper())
        titulo_cel.font = Font(bold=True, size=10)
        for coluna in range(1, 7):
            ws.cell(row=linha, column=coluna).border = Border(top=_BORDA_GROSSA, bottom=_BORDA_GROSSA)
        linha += 1

        for item in categoria.itens:
            linha_inicio = linha
            desc_cel = ws.cell(row=linha, column=2, value=item.descricao or "")
            desc_cel.alignment = Alignment(wrap_text=True, vertical="top")
            desc_cel.font = Font(size=9)

            qtd_cel = ws.cell(row=linha, column=3, value=float(item.quantidade) if item.quantidade else None)
            qtd_cel.number_format = "#,##0.00"
            qtd_cel.alignment = Alignment(horizontal="center", vertical="top")
            qtd_cel.font = Font(size=9)

            un_cel = ws.cell(row=linha, column=4, value=item.unidade or "")
            un_cel.alignment = Alignment(horizontal="center", vertical="top")
            un_cel.font = Font(size=9)

            if item.status_cobranca == models.COBRANCA_NORMAL:
                pu_cel = ws.cell(row=linha, column=5, value=float(item.preco_unitario) if item.preco_unitario is not None else None)
                pu_cel.number_format = '"R$" #,##0.00'
                pu_cel.alignment = Alignment(horizontal="right", vertical="top")
                pu_cel.font = Font(size=9)

                pt_cel = ws.cell(row=linha, column=6, value=float(item.preco_total) if item.preco_total is not None else None)
                pt_cel.number_format = '"R$" #,##0.00'
                pt_cel.alignment = Alignment(horizontal="right", vertical="top")
                pt_cel.font = Font(size=9)
            else:
                ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=6)
                rotulo_cel = ws.cell(row=linha, column=5, value=STATUS_LABEL.get(item.status_cobranca, ""))
                rotulo_cel.alignment = Alignment(horizontal="center", vertical="top")
                rotulo_cel.font = Font(size=9, italic=True)

            for coluna in range(1, 7):
                ws.cell(row=linha, column=coluna).border = Border(bottom=_BORDA_FINA)
            ws.row_dimensions[linha_inicio].height = None
            linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="TOTAL DA PROPOSTA").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=5)
    total_cel = ws.cell(row=linha, column=6, value=float(totais.total_geral))
    total_cel.number_format = '"R$" #,##0.00'
    total_cel.font = Font(bold=True, size=11)
    total_cel.alignment = Alignment(horizontal="right")
    for coluna in range(1, 7):
        ws.cell(row=linha, column=coluna).border = Border(top=_BORDA_GROSSA, bottom=Side(style="double", color=EMPRESA_COR_BARRA))


def _montar_pagina_condicoes(ws: Worksheet, orcamento: models.Orcamento, totais: schemas.OrcamentoTotais) -> None:
    _aplicar_larguras(ws)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"

    linha = _cabecalho_empresa(ws, "F")
    linha += 1

    ws.cell(row=linha, column=1, value=f"PROPOSTA Nº {orcamento.numero_proposta or '—'}").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    linha += 2

    def secao(numero: str, titulo: str) -> int:
        nonlocal linha
        ws.cell(row=linha, column=1, value=numero).font = Font(bold=True, size=10)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
        celula = ws.cell(row=linha, column=2, value=titulo)
        celula.font = Font(bold=True, size=10)
        linha += 1
        return linha

    def texto(valor: str) -> int:
        nonlocal linha
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
        celula = ws.cell(row=linha, column=2, value=valor)
        celula.alignment = Alignment(wrap_text=True, vertical="top")
        celula.font = Font(size=10)
        linha += 1
        return linha

    secao("1.", "PREÇO GLOBAL:")
    valor_extenso = valor_por_extenso(totais.total_geral)
    texto(f"{_fmt_moeda(totais.total_geral)} ({valor_extenso})")
    linha += 1

    secao("2.", "CONDIÇÕES DE PAGAMENTO:")
    texto(orcamento.condicoes_pagamento or "")
    linha += 1

    secao("3.", "PRAZO DE EXECUÇÃO")
    texto("Conforme cronograma da obra")
    linha += 1

    perc_material = float(orcamento.percentual_material)
    perc_servico = float(orcamento.percentual_servico)
    ws.cell(row=linha, column=1, value="4.").font = Font(bold=True, size=10)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
    celula = ws.cell(
        row=linha,
        column=2,
        value=f"PORCENTAGENS: {perc_material:g}% material / {perc_servico:g}% serviço",
    )
    celula.font = Font(bold=True, size=10)
    linha += 2

    ws.cell(row=linha, column=1, value="5.").font = Font(bold=True, size=10)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
    ws.cell(row=linha, column=2, value="DISPOSIÇÕES GERAIS").font = Font(bold=True, size=10)
    linha += 1

    for indice, item in enumerate(DISPOSICOES_GERAIS, start=1):
        ws.cell(row=linha, column=1, value=f"5.{indice}").font = Font(size=9)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
        celula = ws.cell(row=linha, column=2, value=item)
        celula.alignment = Alignment(wrap_text=True, vertical="top")
        celula.font = Font(size=9)
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="DADOS PARA FATURAMENTO:").font = Font(bold=True, size=10)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    linha += 1
    for rotulo, valor in [("Banco:", BANCO), ("Ag.:", AGENCIA), ("Cc.:", CONTA), ("Pix", PIX)]:
        ws.cell(row=linha, column=1, value=f"{rotulo} {valor}").font = Font(size=9)
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
        linha += 1

    linha += 2
    ws.cell(row=linha, column=1, value="Atenciosamente,").font = Font(size=10)
    linha += 2
    ws.cell(row=linha, column=1, value=RESPONSAVEL_NOME).font = Font(bold=True, size=10)
    linha += 1
    ws.cell(row=linha, column=1, value=RESPONSAVEL_CARGO).font = Font(size=9)
    linha += 1
    ws.cell(row=linha, column=1, value=RESPONSAVEL_CONTATO).font = Font(size=9)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)


def gerar_excel(orcamento: models.Orcamento, totais: schemas.OrcamentoTotais) -> BytesIO:
    wb = Workbook()
    ws_itens = wb.active
    ws_itens.title = "Proposta"
    _montar_pagina_itens(ws_itens, orcamento, totais)

    ws_condicoes = wb.create_sheet("Condições")
    _montar_pagina_condicoes(ws_condicoes, orcamento, totais)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
